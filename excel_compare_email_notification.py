import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

old_file = openpyxl.load_workbook("old_file.xlsx", keep_vba=True)
new_file = openpyxl.load_workbook("new_file.xlsx", keep_vba=True)
marker_color = PatternFill(fgColor='00008080', fill_type='solid')

SENDER = 'info@ingo-janssen.de'
PASSWORD = input('E-Mail Account Passwort eingeben: ')
SMTP_SERVER = 'smtp.ionos.de'
SMTP_PORT = 465

RECIPIENT = ['info@ingo-janssen.de']
SUBJECT = 'Ein netter Gruß.'
MESSAGE_TEXT = '''Guten Morgen!
Ich wünsche dir einen wunderschönen Tag!'''
MESSAGE_TEMPLATE = '''
<html>
    <body>
        <p>Hier die Änderungen in den Dateien</p>
        <table>
            <tr>
                <td>Zelle</td>
                <td>Alter Wert</td>
                <td>Neuer Wert</td>
            </tr>
            INHALT
        </table>
    </body>
</html>
'''


def compare_excel_files():
    changes = {}
    for sheetname_old_file, sheetname_new_file in zip(old_file.sheetnames, new_file.sheetnames):
        sheet_old_file = old_file[sheetname_old_file]
        sheet_new_file = new_file[sheetname_new_file]
        # last filled row
        max_row = max(sheet_old_file.max_row, sheet_new_file.max_row)
        # last filled column
        max_column = max(sheet_old_file.max_column, sheet_new_file.max_column)
        print(sheet_old_file, sheet_new_file)
        for col_idx in range(1, max_column + 1):
            for row_idx in range(1, max_row + 1):
                old_cell_value = sheet_old_file.cell(column=col_idx, row=row_idx).value
                new_cell_value = sheet_new_file.cell(column=col_idx, row=row_idx).value
                if old_cell_value != new_cell_value:
                    sheet_new_file.cell(column=col_idx, row=row_idx).fill = marker_color
                    print('\nUnterschied gefunden:')
                    print(f'Spalte: {col_idx}, Zeile: {row_idx}')
                    print(f'Alter Wert: {old_cell_value}')
                    print(f'Neuer Wert: {new_cell_value}')
                    cell_name = f'{get_column_letter(col_idx)}{row_idx}'
                    changes[cell_name] = [old_cell_value, new_cell_value]

            print(f'Spalte: {col_idx} fertig.')

    new_file.save('result.xlsx')
    new_file.close()
    send_email_notification(changes)


def send_email_notification(changes):
    message = MIMEMultipart()
    message['Subject'] = SUBJECT
    message['From'] = SENDER
    message['To'] = ','.join(RECIPIENT)

    table_content = [f'<tr><td>{cell}</td><td>{values[0]}</td><td>{values[1]}</td>' for cell, values in changes.items()]
    message_content = MESSAGE_TEMPLATE.replace('INHALT', ''.join(table_content))

    message_content_html = MIMEText(message_content, 'html')
    try:
        message.attach(message_content_html)
    except (IOError, ValueError) as e:
        print(e)
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as mail_server:
        mail_server.login(SENDER, PASSWORD)
        mail_server.sendmail(SENDER, RECIPIENT, message.as_string())


if __name__ == '__main__':
    compare_excel_files()
